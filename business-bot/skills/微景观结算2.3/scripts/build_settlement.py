#!/usr/bin/env python3
import argparse
import datetime as dt
import json
import os
import re
import shutil
import zipfile
from collections import Counter, defaultdict
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_RULES = {
    "keywords": ["微景观", "苔藓"],
    "commission_rate": "0.006",
    "commission": {
        "exclude_status_keywords": ["已取消", "待付款"],
        "basis": "all_target_orders_except_cancelled_unpaid",
        "include_refund_orders": True,
        "include_pending_orders": True,
    },
    "platform_income": {
        "exclude_status_keywords": ["退款成功", "退款中", "售后处理中", "已取消", "待发货", "待付款"],
        "require_effective_cost_match": True,
        "exclude_zero_goods_cost": True,
        "special_include_orders": [],
        "special_exclude_orders": [],
    },
    "platform_cost": {
        "require_qty_match": True,
        "include_shipping_surcharge": False,
        "require_shipped_status": True,
    },
    "status_rules": {
        "shipped_keywords": ["已发货", "已收货"],
        "refund_success_keywords": ["退款成功"],
        "refunding_keywords": ["退款中", "售后处理中"],
        "cancelled_keywords": ["已取消"],
        "pending_keywords": ["待发货"],
        "unpaid_keywords": ["待付款"],
    },
    "supplier_cleaning": {
        "allow_blank_order_inherit_only_without_independent_traits": True,
        "independent_trait_fields": ["收件地址 必须含省市区", "快递单号", "日期"],
        "skip_express_no_equals": ["合计"],
    },
    "store_normalization": {
        "ad_store_name_map": {
            "小须鲸旗舰店": "小须鲸旗舰店",
            "秘牌宠物用品店": "秘牌小店",
            "秘牌宠物用品旗舰店": "秘牌旗舰店",
            "金鱼水族旗舰店": "金鱼旗舰店"
        },
        "platform_store_name_map": {},
        "supplier_store_name_map": {"秘牌2": "秘牌2店"},
        "supplier_source_file_store_map": {},
        "platform_supplier_store_map": {}
    },
    "summary": {
        "canonical_store_names": ["小须鲸旗舰店", "秘牌旗舰店", "秘牌小店", "金鱼旗舰店"],
    },
    "checks": {
        "strict": True,
        "expected_metrics": {},
        "fail_on_missing_required_input": True,
        "fail_on_qty_mismatch_match_success": True,
        "fail_on_income_check": True,
        "warn_on_real_abnormal_orders": True,
    },
}

THIN = Side(style="thin", color="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="FCE4D6")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
TITLE_FILL = PatternFill("solid", fgColor="0F243E")
FOCUS_RAW_FILL = PatternFill("solid", fgColor="FDE9D9")
FOCUS_MATCH_FILL = PatternFill("solid", fgColor="EAF4E6")
FOCUS_DIFF_FILL = PatternFill("solid", fgColor="EDE7F6")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def deep_merge(base, override):
    if not isinstance(base, dict) or not isinstance(override, dict):
        return override
    merged = dict(base)
    for k, v in override.items():
        if k in merged and isinstance(merged[k], dict) and isinstance(v, dict):
            merged[k] = deep_merge(merged[k], v)
        else:
            merged[k] = v
    return merged


def auto_find_period_rule_file(start_date, end_date, references_dir):
    candidates = [
        os.path.join(references_dir, f'example_rules_{start_date.isoformat()}_{end_date.isoformat()}.json'),
        os.path.join(references_dir, f'rules_{start_date.isoformat()}_{end_date.isoformat()}.json'),
        os.path.join(references_dir, f'batch_rules_{start_date.isoformat()}_{end_date.isoformat()}.json'),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def load_rules(rule_path=None, start_date=None, end_date=None, references_dir=None):
    rules = json.loads(json.dumps(DEFAULT_RULES))
    if rule_path:
        with open(rule_path, 'r', encoding='utf-8') as f:
            user_rules = json.load(f)
        rules = deep_merge(rules, user_rules)

    default_rule_name = 'default_rules.json'
    current_rule_name = os.path.basename(rule_path) if rule_path else default_rule_name
    should_try_period_rules = (current_rule_name == default_rule_name)
    if should_try_period_rules and start_date and end_date and references_dir:
        auto_rule_path = auto_find_period_rule_file(start_date, end_date, references_dir)
        if auto_rule_path and os.path.abspath(auto_rule_path) != os.path.abspath(rule_path or ''):
            with open(auto_rule_path, 'r', encoding='utf-8') as f:
                period_rules = json.load(f)
            rules = deep_merge(rules, period_rules)
            rules.setdefault('_meta', {})['auto_loaded_period_rules'] = auto_rule_path
    return rules


def d(v):
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v).strip())
    except Exception:
        return Decimal("0")


def parse_extra(v):
    # 2026-03-15 固化：供应商“快递附加运费”列按备注处理，不参与任何成本计算。
    # 无论出现“运费+1”“礼盒+3”“这单算30”等文本，统一按 0。
    return Decimal("0")


def is_target_title(text, keywords):
    s = str(text or "")
    return any(k in s for k in keywords)


def safe_name(name):
    name = str(name or '')
    name = re.sub(r'\d{4}\.\d{1,2}\.\d{1,2}-\d{4}\.\d{1,2}\.\d{1,2}', '', name)
    name = re.sub(r'\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}', '', name)
    name = name.replace('.xlsx', '').strip('-_ ')
    return name


def normalize_store_name(name, mapping=None):
    mapping = mapping or {}
    raw = safe_name(str(name or ''))
    return mapping.get(raw, raw)


def parse_period_from_name(name):
    s = str(name or '')
    patterns = [
        r'(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*[-~至到]\s*(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})',
        r'(\d{1,2})[.\-/](\d{1,2})\s*[-~至到]\s*(\d{1,2})[.\-/](\d{1,2})',
    ]
    for pat in patterns:
        m = re.search(pat, s)
        if not m:
            continue
        groups = m.groups()
        if len(groups) == 6:
            return dt.date(int(groups[0]), int(groups[1]), int(groups[2])), dt.date(int(groups[3]), int(groups[4]), int(groups[5]))
        if len(groups) == 4:
            today_year = dt.date.today().year
            return dt.date(today_year, int(groups[0]), int(groups[1])), dt.date(today_year, int(groups[2]), int(groups[3]))
    return None, None


def is_valid_supplier_sheet_name(sheet_name):
    s = str(sheet_name or '').strip()
    if not s:
        return False
    bad_keywords = ['副本', '复制', 'copy', 'Copy', 'COPY']
    return not any(k in s for k in bad_keywords)


def date_in_period(value, start_date=None, end_date=None):
    if isinstance(value, dt.datetime):
        value = value.date()
    if not isinstance(value, dt.date):
        return False
    if start_date and value < start_date:
        return False
    if end_date and value > end_date:
        return False
    return True


def supplier_sheet_period_stats(ws, date_col_idx, qty_col_idx, start_date=None, end_date=None):
    total_qty = Decimal('0')
    in_period_qty = Decimal('0')
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        date_val = row[date_col_idx] if date_col_idx < len(row) else None
        qty = d(row[qty_col_idx] if qty_col_idx < len(row) else 0)
        if qty <= 0 or not isinstance(date_val, dt.datetime):
            continue
        total_qty += qty
        if date_in_period(date_val, start_date, end_date):
            in_period_qty += qty
    return total_qty, in_period_qty


def unpack_zip(zip_path, temp_dir):
    shutil.rmtree(temp_dir, ignore_errors=True)
    os.makedirs(temp_dir, exist_ok=True)
    with zipfile.ZipFile(zip_path) as z:
        z.extractall(temp_dir)
    subdirs = [os.path.join(temp_dir, x) for x in os.listdir(temp_dir)]
    for p in subdirs:
        if os.path.isdir(p):
            return p
    return temp_dir


def open_xlsx(path):
    return load_workbook(path, read_only=True, data_only=True)


def aggregate_platform(rows):
    agg = {}
    seen_exact = set()
    for r in rows:
        # 平台原始文件可能出现“同一订单被错误混入别的店铺文件”的脏数据。
        # 这里去重必须忽略店铺/来源文件，只看订单本身字段，避免跨文件重复累计件数和收入。
        exact_key = (
            str(r.get('order_no') or '').strip(),
            str(r.get('title') or '').strip(),
            str(r.get('spec') or '').strip(),
            str(r.get('status') or '').strip(),
            str(r.get('after_sale_status') or '').strip(),
            str(r.get('merchant_income') or '').strip(),
            str(r.get('qty') or '').strip(),
            str(r.get('ship_time') or '').strip(),
        )
        if exact_key in seen_exact:
            continue
        seen_exact.add(exact_key)
        key = r['order_no']
        if key not in agg:
            agg[key] = dict(r)
            agg[key]['qty'] = d(r['qty'])
            agg[key]['merchant_income'] = d(r['merchant_income'])
        else:
            agg[key]['qty'] += d(r['qty'])
            agg[key]['merchant_income'] += d(r['merchant_income'])
            agg[key]['title'] = f"{agg[key]['title']} | {r['title']}"
    return agg


def aggregate_supplier(rows):
    agg = {}
    for r in rows:
        key = r['order_no']
        if not key:
            continue
        if key not in agg:
            agg[key] = {
                'order_no': key,
                'supplier_store': r['supplier_store'],
                'source_file': r['source_file'],
                'source_sheet': r['source_sheet'],
                'date': r['date'],
                'product_names': [],
                'qty': Decimal('0'),
                'goods_cost': Decimal('0'),
                'shipping_cost': Decimal('0'),
                'total_cost': Decimal('0'),
                'row_count': 0,
            }
        agg[key]['product_names'].append(str(r['product'] or ''))
        agg[key]['qty'] += d(r['qty'])
        agg[key]['goods_cost'] += d(r['goods_cost'])
        agg[key]['shipping_cost'] += d(r['shipping_cost'])
        agg[key]['total_cost'] += d(r['goods_cost']) + d(r['shipping_cost'])
        agg[key]['row_count'] += 1
    for v in agg.values():
        v['product_names'] = ' | '.join([x for x in v['product_names'] if x])
    return agg


def require_columns(header, required, file_name, sheet_name):
    missing = [c for c in required if c not in header]
    if missing:
        raise ValueError(f'{file_name} / {sheet_name} 缺少必需字段: {", ".join(missing)}')


def normalize_supplier_header(header):
    fixed = list(header)
    expected_positions = {
        0: '店铺',
        1: '平台',
        2: '日期',
        3: '订单号',
        4: '收件地址 必须含省市区',
        5: '品名',
        6: '数量',
        7: '备注',
        8: '快递单号',
        9: '代发价格',
        10: '快递附加运费',
        11: '是否结算',
    }
    for idx, expected in expected_positions.items():
        current = fixed[idx] if idx < len(fixed) else ''
        if not current:
            if idx < len(fixed):
                fixed[idx] = expected
        elif current != expected:
            if expected in ('店铺', '订单号', '快递单号'):
                fixed[idx] = expected
    return fixed


def load_platform(base_dir, rules):
    rows = []
    files = []
    diagnostics = []
    keywords = rules['keywords']
    store_mapping = rules.get('store_normalization', {}).get('platform_store_name_map', {})
    required = ['商品', '订单号', '订单状态', '商家实收金额(元)', '商品数量(件)']
    for fn in sorted(os.listdir(base_dir)):
        if not fn.endswith('.xlsx') or '供应商发货' in fn or '广告数据' in fn:
            continue
        files.append(fn)
        wb = open_xlsx(os.path.join(base_dir, fn))
        ws = wb[wb.sheetnames[0]]
        header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        require_columns(header, required, fn, ws.title)
        idx = {k: i for i, k in enumerate(header)}
        raw_store_name = safe_name(fn)
        normalized_store_name = normalize_store_name(raw_store_name, store_mapping)
        if raw_store_name != normalized_store_name:
            diagnostics.append(['平台', fn, ws.title, '店铺名规范化', raw_store_name, normalized_store_name, '按 rules.store_normalization.platform_store_name_map 转换'])
        for row in ws.iter_rows(min_row=2, values_only=True):
            title = row[idx.get('商品', 0)]
            if not is_target_title(title, keywords):
                continue
            rows.append({
                'store': normalized_store_name,
                'raw_store': raw_store_name,
                'source_file': fn,
                'title': str(title or ''),
                'spec': str(row[idx['商品规格']] or '').strip() if '商品规格' in idx else '',
                'order_no': str(row[idx['订单号']] or '').strip(),
                'status': str(row[idx['订单状态']] or '').strip(),
                'after_sale_status': str(row[idx['售后状态']] or '').strip() if '售后状态' in idx else '',
                'merchant_income': row[idx['商家实收金额(元)']],
                'qty': row[idx['商品数量(件)']],
                'ship_time': row[idx['发货时间']] if '发货时间' in idx else None,
            })
    return rows, files, diagnostics


def load_supplier(base_dir, start_date, end_date, rules):
    rows = []
    files = []
    diagnostics = []
    cleaning = rules['supplier_cleaning']
    store_norm = rules.get('store_normalization', {})
    supplier_store_mapping = store_norm.get('supplier_store_name_map', {})
    supplier_source_file_store_map = store_norm.get('supplier_source_file_store_map', {})
    skip_express_values = set(cleaning.get('skip_express_no_equals', []))
    for fn in sorted(os.listdir(base_dir)):
        if '供应商发货' not in fn or not fn.endswith('.xlsx'):
            continue
        files.append(fn)
        wb = open_xlsx(os.path.join(base_dir, fn))
        file_period_start, file_period_end = parse_period_from_name(fn)
        for sheet in wb.sheetnames:
            if not is_valid_supplier_sheet_name(sheet):
                diagnostics.append(['供应商', fn, sheet, '跳过副本/复制sheet', sheet, '跳过', '副本sheet不参与结算，避免历史重复数据污染'])
                continue
            ws = wb[sheet]
            it = ws.iter_rows(min_row=1, values_only=True)
            try:
                raw_header = [str(c).strip() if c is not None else '' for c in next(it)]
            except StopIteration:
                continue
            header = normalize_supplier_header(raw_header)
            if raw_header != header:
                diagnostics.append(['供应商', fn, sheet, '表头纠正', ' | '.join(raw_header[:12]), ' | '.join(header[:12]), '自动纠正非标准供应商表头'])
            if '订单号' not in header:
                continue
            required = ['店铺', '日期', '订单号', '品名', '数量', '代发价格', '快递附加运费']
            require_columns(header, required, fn, sheet)
            col = {k: i for i, k in enumerate(header)}
            if file_period_start and file_period_end and '日期' in col and '数量' in col:
                sheet_total_qty, sheet_in_period_qty = supplier_sheet_period_stats(ws, col['日期'], col['数量'], file_period_start, file_period_end)
                if sheet_total_qty > 0:
                    coverage = (sheet_in_period_qty / sheet_total_qty)
                    if coverage < Decimal('0.5'):
                        diagnostics.append(['供应商', fn, sheet, '跳过低覆盖历史sheet', str(sheet_in_period_qty), str(sheet_total_qty), '跳过', 'sheet内本周期件数覆盖率过低，判定为历史/杂项流水，不纳入本批次'])
                        continue
            it = ws.iter_rows(min_row=2, values_only=True)
            prev_order = ''
            prev_store = ''
            for row_idx, row in enumerate(it, start=2):
                if not any(v is not None and str(v).strip() for v in row):
                    continue
                raw_order = str(row[col['订单号']] or '').strip()
                store = str(row[col['店铺']] or '').strip()
                product = row[col.get('品名')]
                date_val = row[col.get('日期')]
                goods_cost = row[col.get('代发价格')]
                express_no = row[col.get('快递单号')] if '快递单号' in col else None

                if (not raw_order and not isinstance(date_val, dt.datetime) and not str(product or '').strip() and not str(express_no or '').strip()):
                    continue
                if str(express_no or '').strip() in skip_express_values:
                    diagnostics.append(['供应商', fn, sheet, '跳过合计/污染行', f'第{row_idx}行', str(express_no or ''), '跳过', '快递单号命中合计等污染值'])
                    continue

                order = raw_order
                has_independent_traits = False
                if order:
                    prev_order = order
                else:
                    address = row[col.get('收件地址 必须含省市区')] if '收件地址 必须含省市区' in col else None
                    has_independent_traits = bool(str(address or '').strip()) or bool(str(express_no or '').strip()) or isinstance(date_val, dt.datetime)
                    if cleaning.get('exclude_blank_order_rows', False):
                        diagnostics.append(['供应商', fn, sheet, '跳过空订单号行', f'第{row_idx}行', str(product or ''), '跳过', '用户最新口径：空订单号一律不计入结算件数，也不参与匹配'])
                        continue
                    if cleaning.get('allow_blank_order_inherit_only_without_independent_traits', True) and has_independent_traits:
                        diagnostics.append(['供应商', fn, sheet, '无订单号独立发货', f'第{row_idx}行', str(product or ''), '计入发货侧，不参与匹配', '空订单号但带独立特征，视为真实独立发货记录'])
                        order = ''
                    else:
                        if not (str(product or '').strip() or goods_cost not in (None, '')):
                            continue
                        order = prev_order
                if store:
                    prev_store = store
                elif order and prev_order and order == prev_order:
                    # 2026-03-21 固化：只有“同一订单号的补充行”才允许补继上一行店铺。
                    # 空订单号 + 空店铺 + 独立特征（地址/快递单号/日期）的独立发货记录，
                    # 不得因为出现在某个供应商文件里，就偷挂到上一行店铺；应保持空店铺，最终归入“未知店铺”。
                    store = prev_store
                raw_store = store
                normalized_store = normalize_store_name(store, supplier_store_mapping)
                source_store = ''
                source_key = safe_name(fn)
                for key, mapped in supplier_source_file_store_map.items():
                    if safe_name(key) and safe_name(key) in source_key:
                        source_store = mapped
                        break
                if source_store and not normalized_store:
                    diagnostics.append(['供应商', fn, sheet, '来源文件归属补全店铺', raw_store, source_store, '按 rules.store_normalization.supplier_source_file_store_map 补全空/缺失店铺'])
                    normalized_store = source_store
                elif raw_store and normalized_store != raw_store:
                    diagnostics.append(['供应商', fn, sheet, '店铺名规范化', raw_store, normalized_store, '按 rules.store_normalization.supplier_store_name_map 转换'])
                store = normalized_store
                if not order and not has_independent_traits:
                    continue
                if (not str(store or '').strip()) and has_independent_traits and cleaning.get('exclude_blank_store_independent_rows', False):
                    diagnostics.append(['供应商', fn, sheet, '跳过空店铺独立发货', f'第{row_idx}行', str(product or ''), '跳过', '用户最新口径：空店铺独立发货不计入，不挂任何店铺'])
                    continue
                if isinstance(date_val, dt.datetime):
                    # 2026-03-23 修正：供应商明细侧默认不再按“文件名周期/行内日期”二次裁切。
                    # 用户给了哪一批供应商原表，就按这批原表整盘参与匹配与异常明细；
                    # 否则会出现“原始总件数保留了，但供应商有/平台无明细被裁掉”的口径撕裂。
                    pass
                elif raw_order:
                    diagnostics.append(['供应商', fn, sheet, '新订单缺日期', f'第{row_idx}行', raw_order, '跳过', '新订单行没有日期，风险高，已跳过'])
                    continue
                rows.append({
                    'supplier_store': store,
                    'raw_supplier_store': str(row[col['店铺']] or '').strip(),
                    'source_file': fn,
                    'source_sheet': sheet,
                    'date': date_val,
                    'order_no': order,
                    'raw_order_no': raw_order,
                    'product': product,
                    'qty': row[col.get('数量')],
                    'goods_cost': goods_cost,
                    'shipping_cost': parse_extra(row[col.get('快递附加运费')]),
                    'independent_no_order': (not order) and has_independent_traits,
                })
    return rows, files, diagnostics


def load_supplier_raw_totals(base_dir, rules):
    rows = []
    store_norm = rules.get('store_normalization', {})
    supplier_store_mapping = store_norm.get('supplier_store_name_map', {})
    supplier_source_file_store_map = store_norm.get('supplier_source_file_store_map', {})
    for fn in sorted(os.listdir(base_dir)):
        if '供应商发货' not in fn or not fn.endswith('.xlsx'):
            continue
        wb = open_xlsx(os.path.join(base_dir, fn))
        file_period_start, file_period_end = parse_period_from_name(fn)
        for sheet in wb.sheetnames:
            if not is_valid_supplier_sheet_name(sheet):
                continue
            ws = wb[sheet]
            it = ws.iter_rows(min_row=1, values_only=True)
            try:
                raw_header = [str(c).strip() if c is not None else '' for c in next(it)]
            except StopIteration:
                continue
            header = normalize_supplier_header(raw_header)
            if '数量' not in header:
                continue
            col = {k: i for i, k in enumerate(header)}
            if file_period_start and file_period_end and '日期' in col and '数量' in col:
                sheet_total_qty, sheet_in_period_qty = supplier_sheet_period_stats(ws, col['日期'], col['数量'], file_period_start, file_period_end)
                if sheet_total_qty > 0:
                    coverage = (sheet_in_period_qty / sheet_total_qty)
                    if coverage < Decimal('0.5'):
                        continue
            it = ws.iter_rows(min_row=2, values_only=True)
            prev_store = ''
            prev_order = ''
            for row in it:
                if not any(v is not None and str(v).strip() for v in row):
                    continue
                raw_order = str(row[col.get('订单号', 0)] or '').strip()
                raw_store = str(row[col.get('店铺', 0)] or '').strip()
                store = raw_store
                if raw_order:
                    prev_order = raw_order
                if store:
                    prev_store = store
                elif raw_order and prev_order and raw_order == prev_order:
                    # 2026-03-21 固化：供应商原始总件数清洗时，空店铺也只允许在“同一订单号补充行”场景下继承。
                    # 独立无订单号/无店铺的发货记录必须保留为空，后续汇总到“未知店铺”。
                    store = prev_store
                normalized_store = normalize_store_name(store, supplier_store_mapping)
                source_store = ''
                source_key = safe_name(fn)
                for key, mapped in supplier_source_file_store_map.items():
                    if safe_name(key) and safe_name(key) in source_key:
                        source_store = mapped
                        break
                if not normalized_store and source_store:
                    normalized_store = source_store
                store = normalized_store
                express_no = row[col.get('快递单号')] if '快递单号' in col else None
                if str(express_no or '').strip() == '合计':
                    continue
                # 2026-03-16 固化：供应商表格原始总件数 = 用户本次提供的供应商原表有效行件数合计。
                # 这里默认不得按目标周日期、文件名周期、或行内日期再做二次过滤；
                # 否则会把“原始总件数/原始总成本”错误裁小，违背已确认业务口径。
                qty = d(row[col.get('数量')])
                if qty <= 0:
                    continue
                rows.append({
                    'supplier_store': store,
                    'raw_supplier_store': raw_store,
                    'source_store': source_store,
                    'source_file': fn,
                    'source_sheet': sheet,
                    'order_no': raw_order,
                    'date': row[col.get('日期')] if '日期' in col else None,
                    'product': row[col.get('品名')] if '品名' in col else '',
                    'express_no': express_no,
                    'qty': qty,
                })
    return rows


def normalize_ad_store_name(fn, rules):
    name = fn.replace('.xlsx', '')
    name = re.sub(r'\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}广告数据', '', name)
    name = name.replace('广告数据', '').strip()
    mapping = rules.get('store_normalization', {}).get('ad_store_name_map', {})
    return normalize_store_name(name, mapping)


def load_ads(base_dir, rules):
    rows = []
    files = []
    diagnostics = []
    keywords = rules['keywords']
    required = ['商品名称', '总花费(元)', '交易额(元)', '净交易额(元)']
    for fn in sorted(os.listdir(base_dir)):
        if '广告数据' not in fn or not fn.endswith('.xlsx'):
            continue
        files.append(fn)
        wb = open_xlsx(os.path.join(base_dir, fn))
        ws = wb[wb.sheetnames[0]]
        header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        require_columns(header, required, fn, ws.title)
        idx = {k: i for i, k in enumerate(header)}
        raw_store_name = re.sub(r'\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}广告数据', '', fn.replace('.xlsx', '')).replace('广告数据', '').strip()
        store_name = normalize_ad_store_name(fn, rules)
        if safe_name(raw_store_name) != store_name:
            diagnostics.append(['广告', fn, ws.title, '店铺名规范化', safe_name(raw_store_name), store_name, '按 rules.store_normalization.ad_store_name_map 转换'])
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[idx.get('商品名称', 1)] or '').strip()
            if not name or name == '总计' or not is_target_title(name, keywords):
                continue
            rows.append({
                'store': store_name,
                'source_file': fn,
                'product_name': name,
                'spend': d(row[idx.get('总花费(元)')]),
                'deal_amount': d(row[idx.get('交易额(元)')]),
                'net_amount': d(row[idx.get('净交易额(元)')]),
            })
    return rows, files, diagnostics


def infer_store_map(platform_orders, supplier_rows, rules):
    counts = Counter()
    order_to_store = {k: v['store'] for k, v in platform_orders.items()}
    for r in supplier_rows:
        order = r['order_no']
        if order in order_to_store:
            counts[(order_to_store[order], r['supplier_store'])] += 1
    mapping = dict(rules.get('store_normalization', {}).get('platform_supplier_store_map', {}) or {})
    for (pstore, sstore), _ in counts.most_common():
        if pstore not in mapping:
            mapping[pstore] = sstore
    return mapping, counts


def contains_any(text, keywords):
    s = str(text or '')
    return any(k in s for k in keywords)


def status_flags(status, ship_time, rules, after_sale_status=''):
    status_rules = rules['status_rules']
    s = str(status or '')
    refund_success = contains_any(s, status_rules['refund_success_keywords'])
    refunding = contains_any(s, status_rules['refunding_keywords'])
    cancelled = contains_any(s, status_rules['cancelled_keywords'])
    pending = contains_any(s, status_rules['pending_keywords'])
    unpaid = contains_any(s, status_rules.get('unpaid_keywords', []))

    explicit_unshipped = '未发货' in s
    explicit_shipped = contains_any(s, status_rules['shipped_keywords'])
    aftersale_or_refund = refund_success or refunding

    # 平台已发货主看订单状态，不再让发货时间单独触发。
    # 已发货，售后处理中 / 已发货，退款成功 仍算已发货；
    # 已取消、未发货退款成功 一律不算。
    if cancelled:
        shipped = False
    elif explicit_unshipped and aftersale_or_refund:
        shipped = False
    else:
        shipped = explicit_shipped

    unshipped_refund = (explicit_unshipped and aftersale_or_refund) or cancelled
    shipped_refund = aftersale_or_refund and (not explicit_unshipped) and explicit_shipped

    excluded_keywords = rules['platform_income']['exclude_status_keywords']
    valid_income = not contains_any(s, excluded_keywords)
    if unpaid:
        valid_income = False
    # 售后状态独立判断：含退款成功/售后处理中/退款中也剔除
    after_sale_excl_kws = ['退款成功', '售后处理中', '退款中']
    if any(k in str(after_sale_status or '') for k in after_sale_excl_kws):
        valid_income = False
        if not refund_success:
            refund_success = '售后处理中' in str(after_sale_status or '') or '退款成功' in str(after_sale_status or '')

    return {
        'shipped': shipped,
        'refund': refund_success,
        'refunding': refunding,
        'cancelled': cancelled,
        'pending': pending,
        'unpaid': unpaid,
        'unshipped_refund': unshipped_refund,
        'shipped_refund': shipped_refund,
        'valid_income': valid_income,
    }


def is_counted_platform_shipped(rec):
    # 2026-03-21 修正：平台已发货件数只看订单状态，不扣货品成本为0的订单。
    # 货品成本=0限制只用于平台发货总成本计算，不影响件数统计。
    return bool(rec.get('shipped'))



def is_valid_commission_order(status, rules):
    commission_rules = rules.get('commission', {}) or {}
    excluded_keywords = commission_rules.get('exclude_status_keywords', ['已取消', '待付款'])
    s = str(status or '')
    return not contains_any(s, excluded_keywords)


def autosize(ws, widths=None):
    widths = widths or {}
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        if letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
            continue
        max_len = 0
        for c in col_cells:
            try:
                val = '' if c.value is None else str(c.value)
                max_len = max(max_len, min(len(val), 50))
            except Exception:
                pass
        ws.column_dimensions[letter].width = max(12, max_len + 2)


def style_table(ws, header_row=1):
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = LEFT


def apply_row_borders(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fmt_currency(cell):
    cell.number_format = '¥#,##0.00'


def fmt_pct(cell):
    cell.number_format = '0.00%'


def add_summary_sheet(wb, summary_rows, report_period, rules, analysis_lines=None):
    ws = wb.active
    ws.title = '总表'
    ws.merge_cells('A1:S1')
    ws['A1'] = f'微景观结算2.3｜{report_period}'
    ws['A1'].fill = TITLE_FILL
    ws['A1'].font = Font(color='FFFFFF', bold=True, size=14)
    ws['A1'].alignment = CENTER

    total_row = summary_rows[-1]
    cards = [
        ('平台收入', total_row['income']),
        ('平台发货总成本', total_row['platform_cost']),
        ('广告费', total_row['ad_spend']),
        ('平台利润', total_row['profit']),
        ('最终异常订单', total_row.get('real_abnormal_count', 0)),
    ]
    start_col = 1
    for title, value in cards:
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 2)
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col + 2)
        ws.cell(2, start_col, title)
        ws.cell(2, start_col).fill = SUB_FILL
        ws.cell(2, start_col).font = BOLD
        ws.cell(2, start_col).alignment = CENTER
        ws.cell(3, start_col, float(value) if isinstance(value, Decimal) else value)
        ws.cell(3, start_col).alignment = CENTER
        if title in ('平台收入', '平台发货总成本', '广告费', '平台利润'):
            fmt_currency(ws.cell(3, start_col))
        fill = GOOD_FILL if title == '平台利润' and value >= 0 else (BAD_FILL if title == '平台利润' and value < 0 else WARN_FILL)
        ws.cell(3, start_col).fill = fill
        for r in (2, 3):
            for c in range(start_col, start_col + 3):
                ws.cell(r, c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        start_col += 4

    ws['A5'] = '店铺'
    headers = [
        '平台下单件数', '有效销售件数', '平台已发货件数', '供应商表格原始总件数', '供应商平台匹配件数', '供应商与平台发货件数差', '平台发货总成本', '发货表格总成本',
        '平台收入(不含退款)', '佣金基数', '广告费', f'平台佣金({Decimal(rules["commission_rate"]) * 100}%)', '平台利润', '未发货退款件数',
        '未发货退款率', '已发货退款件数', '已发货退款率', '成本率', '广告费率'
    ]
    for i, h in enumerate(headers, start=2):
        ws.cell(5, i, h)
    style_table(ws, 5)
    for col, fill in ((5, FOCUS_RAW_FILL), (6, FOCUS_MATCH_FILL), (7, FOCUS_DIFF_FILL)):
        ws.cell(5, col).fill = fill
        ws.cell(5, col).font = BOLD
        ws.cell(5, col).alignment = CENTER

    row = 6
    for item in summary_rows:
        ws.cell(row, 1, item['store'])
        vals = [
            item['platform_order_qty'], item['valid_sales_qty'], item['shipped_qty'], item['supplier_raw_qty'], item['supplier_qty'], item['ship_qty_diff'], item['platform_cost'], item['supplier_cost'],
            item['income'], item['commission_base'], item['ad_spend'], item['commission'], item['profit'], item['unshipped_refund_qty'],
            item['unshipped_refund_rate'], item['shipped_refund_qty'], item['shipped_refund_rate'], item['cost_rate'], item['ad_rate']
        ]
        for col, v in enumerate(vals, start=2):
            c = ws.cell(row, col, float(v) if isinstance(v, Decimal) else v)
            if col in (8, 9, 10, 11, 12, 13, 14):
                fmt_currency(c)
            if col in (16, 18, 19, 20):
                fmt_pct(c)
        for col, fill in ((5, FOCUS_RAW_FILL), (6, FOCUS_MATCH_FILL), (7, FOCUS_DIFF_FILL)):
            ws.cell(row, col).fill = fill
        if item['store'] == '总计':
            for c in ws[row]:
                c.font = BOLD
            ws.cell(row, 5).fill = FOCUS_RAW_FILL
            ws.cell(row, 6).fill = FOCUS_MATCH_FILL
            ws.cell(row, 7).fill = FOCUS_DIFF_FILL
            for col in [1,2,3,4,8,9,10,11,12,13,14,15,16,17,18,19,20]:
                ws.cell(row, col).fill = SUB_FILL
        row += 1

    notes_row = row + 2
    ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=12)
    note_text = '口径说明：平台收入按规则文件判定，默认仅统计已发货/已收货且非退款成功、非退款中、非售后处理中、非已取消、非待发货、非待付款，且有有效货品成本匹配的订单；佣金基数默认按拼多多相关出单统计，退款单也计佣金，仅剔除已取消、待付款；平台已发货件数默认按物流已发货口径统计，但扣除货品成本为0的订单；平台发货总成本默认只统计匹配成功订单的货品成本，不含快递附加运费；订单号一致但件数不一致不算匹配成功；发货表格总成本为供应商本周相关发货全部成本。'
    ws.cell(notes_row, 1, note_text)
    ws.cell(notes_row, 1).fill = WARN_FILL
    ws.cell(notes_row, 1).alignment = LEFT

    if analysis_lines:
        analysis_title_row = notes_row + 2
        ws.merge_cells(start_row=analysis_title_row, start_column=1, end_row=analysis_title_row, end_column=12)
        ws.cell(analysis_title_row, 1, '运营分析报告')
        ws.cell(analysis_title_row, 1).fill = TITLE_FILL
        ws.cell(analysis_title_row, 1).font = WHITE_FONT
        ws.cell(analysis_title_row, 1).alignment = LEFT
        for i, line in enumerate(analysis_lines, start=analysis_title_row + 1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=12)
            ws.cell(i, 1, line)
            ws.cell(i, 1).alignment = LEFT

    apply_row_borders(ws, 2, 3, 1, 19)
    apply_row_borders(ws, 5, row - 1, 1, 20)
    apply_row_borders(ws, notes_row, notes_row, 1, 12)
    if analysis_lines:
        apply_row_borders(ws, analysis_title_row, analysis_title_row + len(analysis_lines), 1, 12)
    autosize(ws, {'A': 18})
    ws.freeze_panes = 'B4'


def write_rows(ws, headers, rows, currency_cols=None, pct_cols=None, bad_rows=None):
    currency_cols = set(currency_cols or [])
    pct_cols = set(pct_cols or [])
    bad_rows = set(bad_rows or [])
    ws.append(headers)
    for r_idx, row in enumerate(rows, start=2):
        ws.append(row)
        if r_idx in bad_rows:
            for c in ws[r_idx]:
                c.fill = BAD_FILL
    style_table(ws, 1)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in currency_cols:
                fmt_currency(cell)
            if cell.column in pct_cols:
                fmt_pct(cell)
    autosize(ws)
    ws.freeze_panes = 'A2'


def build_validation_rows(summary_rows, platform_detail, supplier_no_platform, platform_no_supplier, platform_no_supplier_pending, platform_no_supplier_real_abnormal, rules, validation_context=None):
    checks = rules['checks']
    expected_metrics = checks.get('expected_metrics', {}) or {}
    total_row = summary_rows[-1]
    validation_rows = []
    validation_context = validation_context or {}

    computed = {
        'total_income': total_row['income'],
        'total_platform_cost': total_row['platform_cost'],
        'total_ad_spend': total_row['ad_spend'],
        'total_profit': total_row['profit'],
        'real_abnormal_count': total_row['real_abnormal_count'],
        'pending_unmatched_count': total_row['pending_unmatched_count'],
        'supplier_only_count': Decimal(str(len(supplier_no_platform))),
        'platform_only_count': Decimal(str(len(platform_no_supplier))),
        'supplier_raw_total_qty': total_row['supplier_raw_qty'],
        'supplier_settlement_total_qty': total_row['supplier_qty'],
    }
    for key, expected in expected_metrics.items():
        actual = computed.get(key)
        if actual is None:
            validation_rows.append([key, '', str(expected), 'WARN', '规则里配置了 expected_metrics，但脚本未识别这个指标'])
            continue
        status = 'PASS' if str(actual) == str(expected) else ('FAIL' if checks.get('fail_on_income_check', True) else 'WARN')
        validation_rows.append([key, str(actual), str(expected), status, '预期值校验'])

    qty_mismatch_rows = [r for r in platform_detail if r['order_exists_in_supplier'] and not r['qty_match']]
    validation_rows.append(['qty_mismatch_orders', str(len(qty_mismatch_rows)), '', 'WARN' if qty_mismatch_rows else 'PASS', '订单号一致但件数不一致，必须人工复核'])

    force_include = set(rules['platform_income'].get('special_include_orders', []))
    force_exclude = set(rules['platform_income'].get('special_exclude_orders', []))
    not_found_force_include = sorted([o for o in force_include if o not in {r['order_no'] for r in platform_detail}])
    not_found_force_exclude = sorted([o for o in force_exclude if o not in {r['order_no'] for r in platform_detail}])
    validation_rows.append(['special_include_orders_not_found', str(len(not_found_force_include)), '', 'WARN' if not_found_force_include else 'PASS', '特例强制计入订单未在本批次出现'])
    validation_rows.append(['special_exclude_orders_not_found', str(len(not_found_force_exclude)), '', 'WARN' if not_found_force_exclude else 'PASS', '特例强制剔除订单未在本批次出现'])

    abnormal_status = 'WARN' if (platform_no_supplier_real_abnormal and checks.get('warn_on_real_abnormal_orders', True)) else 'PASS'
    validation_rows.append(['real_abnormal_orders', str(len(platform_no_supplier_real_abnormal)), '', abnormal_status, '最终异常订单需要人工追查'])
    validation_rows.append(['pending_unmatched_orders', str(len(platform_no_supplier_pending)), '', 'WARN' if platform_no_supplier_pending else 'PASS', '待发货未匹配单独列示，不计入最终异常'])

    raw_total_actual = validation_context.get('supplier_raw_total_actual')
    raw_total_from_summary = validation_context.get('supplier_raw_total_from_summary')
    raw_total_status = 'PASS' if str(raw_total_actual) == str(raw_total_from_summary) else ('FAIL' if checks.get('fail_on_supplier_raw_total_mismatch', True) else 'WARN')
    validation_rows.append(['supplier_raw_total_reconcile', str(raw_total_from_summary), str(raw_total_actual), raw_total_status, '供应商原始总件数必须等于逐文件原始件数合计'])

    unmapped_raw_qty = validation_context.get('unmapped_supplier_raw_qty', Decimal('0'))
    unknown_store_summary_qty = validation_context.get('unknown_store_summary_qty', Decimal('0'))
    unmapped_status = 'PASS' if str(d(unmapped_raw_qty)) == str(d(unknown_store_summary_qty)) else ('FAIL' if checks.get('fail_on_supplier_raw_store_unmapped', True) else 'WARN')
    validation_rows.append(['unmapped_supplier_raw_qty', str(unmapped_raw_qty), str(unknown_store_summary_qty), unmapped_status, '空店铺供应商原始件数必须完整归入“未知店铺”，不允许丢失'])

    summary_store_total_actual = validation_context.get('summary_store_total_actual')
    summary_store_total_expected = validation_context.get('summary_store_total_expected')
    summary_total_status = 'PASS' if str(summary_store_total_actual) == str(summary_store_total_expected) else ('FAIL' if checks.get('fail_on_summary_store_total_mismatch', True) else 'WARN')
    validation_rows.append(['summary_supplier_raw_qty_reconcile', str(summary_store_total_actual), str(summary_store_total_expected), summary_total_status, '各店供应商表格原始总件数之和必须等于供应商原始总件数'])

    missing_platform_store_mappings = validation_context.get('missing_platform_store_mappings', []) or []
    mapping_status = 'PASS' if not missing_platform_store_mappings else ('FAIL' if checks.get('fail_on_missing_platform_supplier_store_map', True) else 'WARN')
    validation_rows.append(['missing_platform_supplier_store_map', str(len(missing_platform_store_mappings)), '0', mapping_status, 'summary 店铺必须有明确 platform_supplier_store_map，禁止只靠自动推断'])

    return validation_rows


def apply_expected_check_failures(validation_rows, rules):
    checks = rules['checks']
    if not checks.get('strict', True):
        return
    failures = [r for r in validation_rows if r[3] == 'FAIL']
    if failures:
        details = '; '.join([f'{r[0]} actual={r[1]} expected={r[2]}' for r in failures])
        raise ValueError(f'复核校验失败: {details}')


def build_report(zip_path, output_path, start_date, end_date, rules):
    temp_dir = os.path.join(os.path.dirname(output_path), '.tmp_weijingguan_build')
    base_dir = unpack_zip(zip_path, temp_dir)

    platform_rows, platform_files, platform_diagnostics = load_platform(base_dir, rules)
    platform_orders = aggregate_platform(platform_rows)
    supplier_rows_all, supplier_files, supplier_diagnostics = load_supplier(base_dir, start_date, end_date, rules)
    supplier_raw_rows_all = load_supplier_raw_totals(base_dir, rules)
    ads_rows, ad_files, ad_diagnostics = load_ads(base_dir, rules)

    input_diagnostics = []
    input_diagnostics.extend(platform_diagnostics)
    input_diagnostics.extend(supplier_diagnostics)
    input_diagnostics.extend(ad_diagnostics)

    store_map, map_counts = infer_store_map(platform_orders, supplier_rows_all, rules)
    explicit_platform_supplier_store_map = dict(rules.get('store_normalization', {}).get('platform_supplier_store_map', {}) or {})
    relevant_supplier_stores = set(store_map.values()) | set(explicit_platform_supplier_store_map.values())
    supplier_rows = [r for r in supplier_rows_all if r['supplier_store'] in relevant_supplier_stores]
    supplier_raw_rows = [r for r in supplier_raw_rows_all if r['supplier_store'] in relevant_supplier_stores]
    supplier_orders = aggregate_supplier([r for r in supplier_rows if r['order_no']])
    supplier_independent_rows = [r for r in supplier_rows if r.get('independent_no_order')]

    platform_store_order_keys = defaultdict(list)
    for order_no, info in platform_orders.items():
        platform_store_order_keys[info['store']].append(order_no)

    ad_by_store = defaultdict(Decimal)
    for r in ads_rows:
        ad_by_store[r['store']] += r['spend']

    commission_rate = Decimal(str(rules['commission_rate']))
    cost_rules = rules['platform_cost']
    income_rules = rules['platform_income']
    special_include_orders = set(income_rules.get('special_include_orders', []))
    special_exclude_orders = set(income_rules.get('special_exclude_orders', []))

    platform_detail = []
    refunds_cost_rows = []
    shipped_refund_rows = []
    duplicate_supplier_shipments = []
    unshipped_refund_with_supplier_cost = []
    platform_no_supplier = []
    for order_no, p in sorted(platform_orders.items()):
        flags = status_flags(p['status'], p['ship_time'], rules, p.get('after_sale_status', ''))
        s = supplier_orders.get(order_no)
        matched_cost = s['total_cost'] if s else Decimal('0')
        goods_cost = s['goods_cost'] if s else Decimal('0')
        shipping_cost = s['shipping_cost'] if s else Decimal('0')
        qty_match = bool(s) and (d(p['qty']) == d(s['qty']))
        effective_match = bool(s)
        if cost_rules.get('require_qty_match', True):
            effective_match = effective_match and qty_match
        platform_cost_value = goods_cost if effective_match else Decimal('0')
        if cost_rules.get('include_shipping_surcharge', False):
            platform_cost_value += shipping_cost if effective_match else Decimal('0')
        rec = {
            'store': p['store'], 'order_no': order_no, 'title': p['title'], 'spec': p.get('spec', ''), 'status': p['status'], 'qty': p['qty'],
            'merchant_income': p['merchant_income'], 'ship_time': p['ship_time'], 'matched': effective_match,
            'order_exists_in_supplier': bool(s),
            'qty_match': qty_match,
            'supplier_store': s['supplier_store'] if s else '', 'supplier_qty': s['qty'] if s else Decimal('0'),
            'goods_cost': platform_cost_value,
            'shipping_cost': shipping_cost if effective_match and cost_rules.get('include_shipping_surcharge', False) else Decimal('0'),
            'matched_cost': platform_cost_value,
            'raw_goods_cost': goods_cost,
            'raw_shipping_cost': shipping_cost,
            'raw_matched_cost': matched_cost,
            **flags,
        }
        platform_detail.append(rec)

        exclude_for_income = False
        if flags['valid_income'] and income_rules.get('require_effective_cost_match', True) and not effective_match:
            exclude_for_income = True
        if flags['valid_income'] and income_rules.get('exclude_zero_goods_cost', True) and goods_cost <= 0:
            exclude_for_income = True
        if order_no in special_exclude_orders:
            exclude_for_income = True
        if order_no in special_include_orders:
            exclude_for_income = False
        rec['exclude_for_income'] = exclude_for_income
        rec['valid_for_commission'] = is_valid_commission_order(p['status'], rules)
        rec['order_commission'] = d(p['merchant_income']) * commission_rate if rec['valid_for_commission'] else Decimal('0')
        rec['order_profit'] = d(p['merchant_income']) - d(platform_cost_value) - rec['order_commission']
        rec['order_profit_rate'] = (rec['order_profit'] / d(p['merchant_income'])) if d(p['merchant_income']) else Decimal('0')

        if not effective_match:
            platform_no_supplier.append(rec)
        if flags['unshipped_refund'] and matched_cost > 0:
            refunds_cost_rows.append(rec)
        if flags['shipped_refund']:
            shipped_refund_rows.append(rec)
        if s and d(s['qty']) > d(p['qty']):
            duplicate_supplier_shipments.append(rec)
        if ('未发货' in str(p['status'] or '')) and ('退款成功' in str(p['status'] or '') or '退款成功' in str(p.get('after_sale_status', '') or '')) and matched_cost > 0:
            unshipped_refund_with_supplier_cost.append(rec)

    supplier_no_platform = []
    for order_no, s in sorted(supplier_orders.items()):
        if order_no not in platform_orders:
            supplier_no_platform.append({
                'supplier_store': s['supplier_store'],
                'order_no': s['order_no'],
                'date': s['date'],
                'product_names': s['product_names'],
                'qty': s['qty'],
                'total_cost': s['total_cost'],
                'source_file': s.get('source_file', ''),
                'source_sheet': s.get('source_sheet', ''),
                'is_independent_no_order': False,
            })
    for r in supplier_independent_rows:
        supplier_no_platform.append({
            'supplier_store': r['supplier_store'],
            'order_no': '',
            'date': r['date'],
            'product_names': str(r['product'] or ''),
            'qty': d(r['qty']),
            'total_cost': d(r['goods_cost']) + d(r['shipping_cost']),
            'source_file': r.get('source_file', ''),
            'source_sheet': r.get('source_sheet', ''),
            'is_independent_no_order': True,
        })

    platform_no_supplier_refund_cancel = [r for r in platform_no_supplier if r['unshipped_refund'] or r['refunding'] or r['cancelled'] or r['unpaid']]
    platform_no_supplier_abnormal = [r for r in platform_no_supplier if (not r['unshipped_refund']) and (not r['refunding']) and (not r['cancelled']) and (not r['unpaid'])]
    platform_no_supplier_pending = [r for r in platform_no_supplier_abnormal if r['pending']]
    platform_no_supplier_real_abnormal = [r for r in platform_no_supplier_abnormal if (not r['pending']) and r['shipped']]

    summary_store_names = rules['summary'].get('canonical_store_names', [])
    platform_name_map = {}
    for pstore in platform_store_order_keys.keys():
        short = normalize_store_name(pstore, rules.get('store_normalization', {}).get('platform_store_name_map', {}))
        platform_name_map[short] = pstore
    canonical = []
    for name in summary_store_names:
        canonical.append((name, platform_name_map.get(name)))
    extra_platform_names = [x for x in platform_name_map.keys() if x not in summary_store_names]
    canonical.extend([(name, platform_name_map.get(name)) for name in extra_platform_names])

    summary_rows = []
    total = defaultdict(Decimal)
    reverse_platform_supplier_map = {k: v for k, v in (rules.get('store_normalization', {}).get('platform_supplier_store_map', {}) or {}).items()}
    for display_name, raw_store in canonical:
        is_unknown_store_row = (display_name == '未知店铺')
        p_details = [r for r in platform_detail if r['store'] == raw_store] if raw_store else []
        mapped_supplier_store = ''
        if raw_store:
            mapped_supplier_store = store_map.get(raw_store, '')
        if not mapped_supplier_store:
            mapped_supplier_store = reverse_platform_supplier_map.get(display_name, '')

        if is_unknown_store_row:
            s_orders = [v for v in supplier_orders.values() if not str(v.get('supplier_store') or '').strip()]
            s_independent_rows = [v for v in supplier_independent_rows if not str(v.get('supplier_store') or '').strip()]
            supplier_raw_qty = sum((r['qty'] for r in supplier_raw_rows if not str(r.get('supplier_store') or '').strip()), Decimal('0'))
        else:
            s_orders = [v for v in supplier_orders.values() if v['supplier_store'] == mapped_supplier_store] if mapped_supplier_store else []
            s_independent_rows = [v for v in supplier_independent_rows if v['supplier_store'] == mapped_supplier_store] if mapped_supplier_store else []
            supplier_raw_qty = sum((r['qty'] for r in supplier_raw_rows if r['supplier_store'] == mapped_supplier_store), Decimal('0')) if mapped_supplier_store else Decimal('0')

        income = sum((r['merchant_income'] for r in p_details if r['valid_income'] and (not r['exclude_for_income'])), Decimal('0'))
        commission_base = sum((r['merchant_income'] for r in p_details if r.get('valid_for_commission')), Decimal('0'))
        platform_order_qty = sum((r['qty'] for r in p_details), Decimal('0'))
        valid_sales_qty = sum((r['qty'] for r in p_details if r['valid_income'] and (not r['exclude_for_income'])), Decimal('0'))
        shipped_qty = sum((r['qty'] for r in p_details if is_counted_platform_shipped(r)), Decimal('0'))
        matched_supplier_orders = [r for r in s_orders if r.get('order_no') in platform_orders]
        supplier_qty = sum((r['qty'] for r in matched_supplier_orders), Decimal('0'))
        platform_cost = sum((r['goods_cost'] for r in p_details if (r['shipped'] or not cost_rules.get('require_shipped_status', True)) and r['matched']), Decimal('0'))
        supplier_cost = sum((r['total_cost'] for r in s_orders), Decimal('0')) + sum((d(r['goods_cost']) + d(r['shipping_cost']) for r in s_independent_rows), Decimal('0'))
        ad_spend = ad_by_store.get(display_name, Decimal('0')) or ad_by_store.get(raw_store or '', Decimal('0'))
        commission = commission_base * commission_rate
        profit = income - platform_cost - commission - ad_spend
        unshipped_refund_qty = sum((r['qty'] for r in p_details if r['unshipped_refund']), Decimal('0'))
        shipped_refund_qty = sum((r['qty'] for r in p_details if r['shipped_refund']), Decimal('0'))
        denom = platform_order_qty if platform_order_qty else Decimal('0')
        row = {
            'store': display_name,
            'platform_order_qty': platform_order_qty,
            'valid_sales_qty': valid_sales_qty,
            'shipped_qty': shipped_qty,
            'supplier_raw_qty': supplier_raw_qty,
            'supplier_qty': supplier_qty,
            'ship_qty_diff': supplier_raw_qty - supplier_qty,
            'platform_cost': platform_cost,
            'supplier_cost': supplier_cost,
            'income': income,
            'commission_base': commission_base,
            'ad_spend': ad_spend,
            'commission': commission,
            'profit': profit,
            'unshipped_refund_qty': unshipped_refund_qty,
            'unshipped_refund_rate': (unshipped_refund_qty / denom) if denom else Decimal('0'),
            'shipped_refund_qty': shipped_refund_qty,
            'shipped_refund_rate': (shipped_refund_qty / denom) if denom else Decimal('0'),
            'cost_rate': (platform_cost / income) if income else Decimal('0'),
            'ad_rate': (ad_spend / income) if income else Decimal('0'),
        }
        summary_rows.append(row)
        for k, v in row.items():
            if k == 'store':
                continue
            total[k] += v
    total_income = total['income']
    total['store'] = '总计'
    total['real_abnormal_count'] = Decimal(str(len(platform_no_supplier_real_abnormal)))
    total['pending_unmatched_count'] = Decimal(str(len(platform_no_supplier_pending)))
    total['refund_cancel_unmatched_count'] = Decimal(str(len(platform_no_supplier_refund_cancel)))
    total['unshipped_refund_rate'] = total['unshipped_refund_qty'] / total['platform_order_qty'] if total['platform_order_qty'] else Decimal('0')
    total['shipped_refund_rate'] = total['shipped_refund_qty'] / total['platform_order_qty'] if total['platform_order_qty'] else Decimal('0')
    total['cost_rate'] = total['platform_cost'] / total_income if total_income else Decimal('0')
    total['ad_rate'] = total['ad_spend'] / total_income if total_income else Decimal('0')
    summary_rows.append(total)

    supplier_raw_total_actual = sum((r['qty'] for r in supplier_raw_rows_all), Decimal('0'))
    supplier_raw_total_from_summary = total['supplier_raw_qty']
    summary_store_total_actual = sum((r['supplier_raw_qty'] for r in summary_rows[:-1]), Decimal('0'))
    summary_store_total_expected = total['supplier_raw_qty']
    unmapped_supplier_raw_qty = sum((r['qty'] for r in supplier_raw_rows_all if not str(r.get('supplier_store') or '').strip()), Decimal('0'))
    missing_platform_store_mappings = []
    for display_name, raw_store in canonical:
        if not raw_store:
            continue
        if raw_store not in store_map and display_name not in explicit_platform_supplier_store_map:
            missing_platform_store_mappings.append(display_name)
    unknown_store_row = next((r for r in summary_rows[:-1] if r.get('store') == '未知店铺'), None)
    unknown_store_summary_qty = unknown_store_row['supplier_raw_qty'] if unknown_store_row else Decimal('0')
    validation_context = {
        'supplier_raw_total_actual': supplier_raw_total_actual,
        'supplier_raw_total_from_summary': supplier_raw_total_from_summary,
        'summary_store_total_actual': summary_store_total_actual,
        'summary_store_total_expected': summary_store_total_expected,
        'unmapped_supplier_raw_qty': unmapped_supplier_raw_qty,
        'unknown_store_summary_qty': unknown_store_summary_qty,
        'missing_platform_store_mappings': missing_platform_store_mappings,
    }

    validation_rows = build_validation_rows(
        summary_rows,
        platform_detail,
        supplier_no_platform,
        platform_no_supplier,
        platform_no_supplier_pending,
        platform_no_supplier_real_abnormal,
        rules,
        validation_context,
    )
    apply_expected_check_failures(validation_rows, rules)

    total_row = summary_rows[-1]
    best_store = max(summary_rows[:-1], key=lambda x: x['profit']) if summary_rows[:-1] else None
    worst_store = min(summary_rows[:-1], key=lambda x: x['profit']) if summary_rows[:-1] else None
    report_lines = [
        f'1. 本周微景观平台下单 {float(total_row["platform_order_qty"]):.0f} 件，有效销售 {float(total_row["valid_sales_qty"]):.0f} 件，平台已发货 {float(total_row["shipped_qty"]):.0f} 件（已扣除货品成本为0订单），供应商表格原始总件数 {float(total_row["supplier_raw_qty"]):.0f} 件，供应商平台匹配件数 {float(total_row["supplier_qty"]):.0f} 件，已发货差异 {float(total_row["ship_qty_diff"]):.0f} 件。',
        f'2. 平台收入 ¥{float(total_row["income"]):,.2f}，佣金基数 ¥{float(total_row["commission_base"]):,.2f}，平台发货总成本 ¥{float(total_row["platform_cost"]):,.2f}，发货表格总成本 ¥{float(total_row["supplier_cost"]):,.2f}，广告费 ¥{float(total_row["ad_spend"]):,.2f}，佣金 ¥{float(total_row["commission"]):,.2f}，最终利润 ¥{float(total_row["profit"]):,.2f}。',
        f'3. 未发货退款 {float(total_row["unshipped_refund_qty"]):.0f} 件，退款率 {float(total_row["unshipped_refund_rate"]):.2%}；已发货退款 {float(total_row["shipped_refund_qty"]):.0f} 件，退款率 {float(total_row["shipped_refund_rate"]):.2%}。',
        f'4. 成本率 {float(total_row["cost_rate"]):.2%}，广告费率 {float(total_row["ad_rate"]):.2%}。',
    ]
    if best_store:
        report_lines.append(f'5. 利润表现最好的是 {best_store["store"]}，利润 ¥{float(best_store["profit"]):,.2f}。')
    if worst_store:
        report_lines.append(f'6. 利润表现最弱的是 {worst_store["store"]}，利润 ¥{float(worst_store["profit"]):,.2f}。')
    report_lines.append('7. 本次口径由规则文件驱动：收入剔除条件、成本计入口径、特例订单、校验指标均可配置。')
    report_lines.append(f'8. 复核提醒：供应商有平台无 {len(supplier_no_platform)} 单；平台有供应商无合计 {len(platform_no_supplier)} 单，其中待发货未匹配 {len(platform_no_supplier_pending)} 单，最终应发未匹配异常 {len(platform_no_supplier_real_abnormal)} 单。')
    report_lines.append('9. 订单号一致但件数不一致不再视为匹配成功，统一回落异常/待核对口径。')
    report_lines.append('10. 交付前必须打开“复核校验”“最终异常订单清单”“订单匹配表”三张表做最后核对。')

    wb = Workbook()
    period_text = f'{start_date.isoformat()} 至 {end_date.isoformat()}'
    add_summary_sheet(wb, summary_rows, period_text, rules, analysis_lines=report_lines)

    ws = wb.create_sheet('老板汇报版')
    ws.merge_cells('A1:H1')
    ws['A1'] = f'微景观周度经营汇报｜{period_text}'
    ws['A1'].fill = TITLE_FILL
    ws['A1'].font = Font(color='FFFFFF', bold=True, size=16)
    ws['A1'].alignment = CENTER
    total_row = summary_rows[-1]
    report_cards = [
        ('平台收入', total_row['income']),
        ('平台发货总成本', total_row['platform_cost']),
        ('广告费', total_row['ad_spend']),
        ('平台利润', total_row['profit']),
        ('最终异常订单', total_row['real_abnormal_count']),
        ('待发货未匹配', total_row['pending_unmatched_count']),
    ]
    col_positions = [1, 4, 7]
    row_positions = [2, 5]
    idx = 0
    for rr in row_positions:
        for cc in col_positions:
            if idx >= len(report_cards):
                break
            title, value = report_cards[idx]
            ws.merge_cells(start_row=rr, start_column=cc, end_row=rr, end_column=cc + 1)
            ws.merge_cells(start_row=rr + 1, start_column=cc, end_row=rr + 1, end_column=cc + 1)
            ws.cell(rr, cc, title)
            ws.cell(rr, cc).fill = SUB_FILL
            ws.cell(rr, cc).font = BOLD
            ws.cell(rr, cc).alignment = CENTER
            ws.cell(rr + 1, cc, float(value) if isinstance(value, Decimal) else value)
            ws.cell(rr + 1, cc).alignment = CENTER
            if title in ('平台收入', '平台发货总成本', '广告费', '平台利润'):
                fmt_currency(ws.cell(rr + 1, cc))
            if title == '平台利润':
                ws.cell(rr + 1, cc).fill = GOOD_FILL if value >= 0 else BAD_FILL
            else:
                ws.cell(rr + 1, cc).fill = WARN_FILL
            for r in (rr, rr + 1):
                for c in range(cc, cc + 2):
                    ws.cell(r, c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            idx += 1
    ws['A9'] = '老板摘要'
    ws['A10'].font = BOLD
    summary_lines = [
        f'本周微景观平台收入 ¥{float(total_row["income"]):,.2f}，平台利润 ¥{float(total_row["profit"]):,.2f}。',
        f'平台已发货 {float(total_row["shipped_qty"]):.0f} 件；供应商表格原始总件数 {float(total_row["supplier_raw_qty"]):.0f} 件；供应商平台匹配件数 {float(total_row["supplier_qty"]):.0f} 件；供应商与平台发货件数差 {float(total_row["ship_qty_diff"]):.0f} 件。',
        f'真正应发未匹配异常 {int(total_row["real_abnormal_count"])} 单，待发货未匹配 {int(total_row["pending_unmatched_count"])} 单。',
        '摘要文案基于规则文件自动生成，交付前仍需人工复核异常与特例订单。',
    ]
    for i, line in enumerate(summary_lines, start=10):
        ws[f'A{i}'] = line
    ws['A11'].fill = FOCUS_DIFF_FILL
    ws['A11'].font = BOLD
    ws['A15'] = '店铺利润排名'
    ws['A15'].font = BOLD
    apply_row_borders(ws, 2, 6, 1, 8)
    apply_row_borders(ws, 10, 13, 1, 8)
    rank_headers = ['店铺', '利润', '收入', '广告费', '最终异常单']
    for i, h in enumerate(rank_headers, start=1):
        ws.cell(16, i, h)
    style_table(ws, 16)
    rank_rows = sorted(summary_rows[:-1], key=lambda x: x['profit'], reverse=True)
    for ridx, item in enumerate(rank_rows, start=17):
        ws.cell(ridx, 1, item['store'])
        ws.cell(ridx, 2, float(item['profit']))
        ws.cell(ridx, 3, float(item['income']))
        ws.cell(ridx, 4, float(item['ad_spend']))
        abnormal = sum(1 for r in platform_no_supplier_real_abnormal if r['store'] == item['store'])
        ws.cell(ridx, 5, abnormal)
        fmt_currency(ws.cell(ridx, 2))
        fmt_currency(ws.cell(ridx, 3))
        fmt_currency(ws.cell(ridx, 4))
    autosize(ws, {'A': 18})
    ws.freeze_panes = 'A16'

    ws = wb.create_sheet('供应商有、平台无')
    # 2026-03-23 固化：此表中的总成本按供应商原表原样展示；
    # 若原表代发价格为负数，这里就直接显示负数，不做绝对值转换。
    rows = [[
        r['supplier_store'],
        r['order_no'],
        r['date'].strftime('%Y-%m-%d') if isinstance(r['date'], dt.datetime) else '',
        r['product_names'],
        float(r['qty']),
        float(r['total_cost']),
        '是' if r.get('is_independent_no_order') else '否',
        r.get('source_file', ''),
        r.get('source_sheet', ''),
    ] for r in supplier_no_platform]
    write_rows(ws, ['供应商店铺', '订单号', '日期', '品名汇总', '发货件数', '总成本', '是否无订单号独立发货', '来源文件', '来源Sheet'], rows, currency_cols={6})

    ws = wb.create_sheet('待发货未匹配')
    rows = [[r['store'], r['order_no'], r['title'], r['status'], float(r['qty']), float(r['merchant_income'])] for r in platform_no_supplier_pending]
    write_rows(ws, ['平台店铺', '订单号', '商品标题', '订单状态', '平台件数', '商家实收金额'], rows, currency_cols={6})

    ws = wb.create_sheet('平台有、供应商无-应发异常')
    rows = [[r['store'], r['order_no'], r['title'], r['status'], float(r['qty']), float(r['merchant_income'])] for r in platform_no_supplier_real_abnormal]
    write_rows(ws, ['平台店铺', '订单号', '商品标题', '订单状态', '平台件数', '商家实收金额'], rows, currency_cols={6})

    ws = wb.create_sheet('最终异常订单清单')
    rows = []
    reverse_store_map = {}
    for p_name, s_name in (rules.get('store_normalization', {}).get('platform_supplier_store_map', {}) or {}).items():
        if s_name:
            reverse_store_map[s_name] = p_name
    diff_detail_rows = []
    for r in supplier_raw_rows_all:
        supplier_store = str(r.get('supplier_store') or '')
        order_no = str(r.get('order_no') or '')
        display_store = reverse_store_map.get(supplier_store, supplier_store) if supplier_store else ''
        reason = ''
        counted = False
        if order_no and supplier_store and (order_no in platform_orders):
            counted = True
        if counted:
            continue
        reasons = []
        if not order_no:
            reasons.append('空订单号，不计入供应商平台匹配件数')
        if not supplier_store:
            reasons.append('空店铺，不挂任何店铺，也不计入供应商平台匹配件数')
        if order_no and order_no not in platform_orders:
            reasons.append('供应商订单号匹配不上平台订单号，不计入供应商平台匹配件数')
        if not reasons:
            reasons.append('未进入供应商平台匹配件数口径')
        diff_detail_rows.append({
            'display_store': display_store,
            'supplier_store': supplier_store,
            'order_no': order_no,
            'date': r.get('date'),
            'product': str(r.get('product') or ''),
            'qty': d(r.get('qty', 0)),
            'cost': Decimal('0'),
            'reason': '；'.join(reasons),
        })
    store_order = {name: i for i, name in enumerate(rules['summary'].get('canonical_store_names', []))}
    diff_detail_rows.sort(key=lambda x: (store_order.get(x['display_store'], 999), str(x['date'] or ''), str(x['order_no'] or ''), str(x['product'] or '')))
    seq = 1
    for item in diff_detail_rows:
        rows.append([
            seq,
            item['display_store'],
            item['supplier_store'],
            item['order_no'],
            item['date'].strftime('%Y-%m-%d') if hasattr(item.get('date'), 'strftime') else '',
            item['product'],
            float(item['qty']),
            float(item['cost']),
            item['reason'],
        ])
        seq += 1
    write_rows(ws, ['序号', '归属店铺', '供应商店铺', '订单号', '日期', '商品/品名', '件数', '供应商总成本', '原因备注'], rows, currency_cols={8})

    ws = wb.create_sheet('未发货退款有成本')
    rows = [[r['store'], r['order_no'], r['title'], r['status'], float(r['qty']), float(r['raw_matched_cost']), r['supplier_store']] for r in refunds_cost_rows]
    write_rows(ws, ['平台店铺', '订单号', '商品标题', '订单状态', '件数', '已匹配原始成本', '供应商店铺'], rows, currency_cols={6})

    ws = wb.create_sheet('重复发货')
    rows = [[r['store'], r['order_no'], r['title'], r['status'], float(r['qty']), float(r['supplier_qty']), float(r['raw_goods_cost']), float(r['raw_shipping_cost']), float(r['raw_matched_cost']), r['supplier_store']] for r in duplicate_supplier_shipments]
    write_rows(ws, ['平台店铺', '订单号', '商品标题', '订单状态', '平台件数', '供应商件数', '供应商货品成本', '供应商快递附加运费', '供应商总成本', '供应商店铺'], rows, currency_cols={7, 8, 9})

    ws = wb.create_sheet('已发货退款明细')
    rows = [[r['store'], r['order_no'], r['title'], r['status'], float(r['qty']), float(r['merchant_income']), float(r['raw_matched_cost'])] for r in shipped_refund_rows]
    write_rows(ws, ['平台店铺', '订单号', '商品标题', '订单状态', '件数', '商家实收金额', '已匹配原始成本'], rows, currency_cols={6, 7})

    ws = wb.create_sheet('利润率低于17%的订单')
    low_margin_rows = []
    for r in platform_detail:
        if d(r['merchant_income']) <= 0:
            continue
        if r['order_profit_rate'] < Decimal('0.17'):
            low_margin_rows.append([
                r['store'],
                r['order_no'],
                r['title'],
                r.get('spec', ''),
                r['status'],
                float(r['qty']),
                float(d(r['merchant_income'])),
                float(d(r['goods_cost'])),
                float(d(r['order_commission'])),
                float(d(r['order_profit'])),
                float(d(r['order_profit_rate'])),
            ])
    write_rows(ws,
        ['平台店铺', '订单号', '商品标题', '商品规格', '订单状态', '件数', '订单收入', '货品成本', '订单佣金', '订单利润', '订单利润率'],
        low_margin_rows, currency_cols={7, 8, 9, 10}, pct_cols={11})

    ws = wb.create_sheet('复核校验')
    rows = [[r[0], r[1], r[2], r[3], r[4]] for r in validation_rows]
    bad_rows = {idx for idx, r in enumerate(validation_rows, start=2) if r[3] in ('FAIL', 'WARN')}
    write_rows(ws, ['校验项', '实际值', '预期值', '状态', '说明'], rows, bad_rows=bad_rows)

    ws = wb.create_sheet('小须鲸专项核对')
    ws['A1'] = '小须鲸专项核对'
    ws['A1'].fill = TITLE_FILL
    ws['A1'].font = WHITE_FONT
    ws['A1'].alignment = CENTER
    ws.merge_cells('A1:T1')
    whale_summary = next((r for r in summary_rows if r.get('store') == '小须鲸旗舰店'), None)
    if whale_summary is None:
        whale_summary = {
            'store': '小须鲸旗舰店',
            'platform_order_qty': Decimal('0'), 'valid_sales_qty': Decimal('0'), 'shipped_qty': Decimal('0'),
            'supplier_raw_qty': Decimal('0'), 'supplier_qty': Decimal('0'), 'ship_qty_diff': Decimal('0'),
            'platform_cost': Decimal('0'), 'supplier_cost': Decimal('0'), 'income': Decimal('0'),
            'commission_base': Decimal('0'), 'ad_spend': Decimal('0'), 'commission': Decimal('0'), 'profit': Decimal('0'),
            'unshipped_refund_qty': Decimal('0'), 'unshipped_refund_rate': Decimal('0'),
            'shipped_refund_qty': Decimal('0'), 'shipped_refund_rate': Decimal('0'),
            'cost_rate': Decimal('0'), 'ad_rate': Decimal('0'),
        }
    ws['A3'] = '说明'
    ws['B3'] = '按总表同口径，仅保留小须鲸旗舰店核心指标，不再展示原专项旧明细。'
    ws['A3'].fill = SUB_FILL
    ws['A3'].font = BOLD
    headers = [
        '店铺', '平台下单件数', '有效销售件数', '平台已发货件数', '供应商表格原始总件数', '供应商平台匹配件数', '供应商与平台发货件数差', '平台发货总成本', '发货表格总成本',
        '平台收入(不含退款)', '佣金基数', '广告费', f'平台佣金({Decimal(rules["commission_rate"]) * 100}%)', '平台利润', '未发货退款件数',
        '未发货退款率', '已发货退款件数', '已发货退款率', '成本率', '广告费率'
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(5, i, h)
    style_table(ws, 5)
    for col, fill in ((5, FOCUS_RAW_FILL), (6, FOCUS_MATCH_FILL), (7, FOCUS_DIFF_FILL)):
        ws.cell(5, col).fill = fill
        ws.cell(5, col).font = BOLD
        ws.cell(5, col).alignment = CENTER
    vals = [
        whale_summary['store'], whale_summary['platform_order_qty'], whale_summary['valid_sales_qty'], whale_summary['shipped_qty'], whale_summary['supplier_raw_qty'], whale_summary['supplier_qty'], whale_summary['ship_qty_diff'], whale_summary['platform_cost'], whale_summary['supplier_cost'],
        whale_summary['income'], whale_summary['commission_base'], whale_summary['ad_spend'], whale_summary['commission'], whale_summary['profit'], whale_summary['unshipped_refund_qty'],
        whale_summary['unshipped_refund_rate'], whale_summary['shipped_refund_qty'], whale_summary['shipped_refund_rate'], whale_summary['cost_rate'], whale_summary['ad_rate']
    ]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(6, col, float(v) if isinstance(v, Decimal) else v)
        if col in (8, 9, 10, 11, 12, 13, 14):
            fmt_currency(c)
        if col in (16, 18, 19, 20):
            fmt_pct(c)
    for col, fill in ((5, FOCUS_RAW_FILL), (6, FOCUS_MATCH_FILL), (7, FOCUS_DIFF_FILL)):
        ws.cell(6, col).fill = fill
    apply_row_borders(ws, 3, 3, 1, 8)
    apply_row_borders(ws, 5, 6, 1, 20)
    autosize(ws, {'A': 18})

    ws = wb.create_sheet('平台微景观订单明细')
    rows = []
    bad = []
    for idx, r in enumerate(platform_detail, start=2):
        if (not r['matched']) or r['refund']:
            bad.append(idx)
        rows.append([
            r['store'], r['order_no'], r['title'], r['status'], float(r['qty']), float(r['merchant_income']),
            '是' if r['shipped'] else '否', '是' if r['matched'] else '否', r['supplier_store'], float(r['supplier_qty']),
            float(r['goods_cost']), float(r['raw_shipping_cost']), float(r['matched_cost']), '是' if r['exclude_for_income'] else '否'
        ])
    write_rows(ws,
        ['店铺', '订单号', '商品标题', '订单状态', '平台件数', '商家实收金额', '是否已发货', '是否匹配供应商', '供应商店铺', '供应商件数', '计入口径货品成本', '原始快递附加运费', '计入口径匹配成本', '是否剔除收入'],
        rows, currency_cols={6, 11, 12, 13}, bad_rows=bad)

    ws = wb.create_sheet('供应商发货明细')
    rows = []
    bad = []
    for idx, (order_no, r) in enumerate(sorted(supplier_orders.items()), start=2):
        if order_no not in platform_orders:
            bad.append(idx)
        rows.append([
            r['supplier_store'], order_no, r['date'].strftime('%Y-%m-%d') if isinstance(r['date'], dt.datetime) else '',
            r['product_names'], float(r['qty']), float(r['goods_cost']), float(r['shipping_cost']), float(r['total_cost']),
            '是' if order_no in platform_orders else '否', r['source_file'], r['source_sheet']
        ])
    write_rows(ws,
        ['供应商店铺', '订单号', '日期', '品名汇总', '发货件数', '货品成本', '快递附加运费', '总成本', '平台是否存在', '来源文件', '来源Sheet'],
        rows, currency_cols={6, 7, 8}, bad_rows=bad)

    ws = wb.create_sheet('订单匹配表')
    rows = []
    for r in platform_detail:
        s = supplier_orders.get(r['order_no'])
        match_result = '匹配成功' if r['matched'] else ('件数不一致' if r['order_exists_in_supplier'] else '平台有、供应商无')
        rows.append([
            r['store'], r['order_no'], float(r['qty']), r['status'], match_result,
            s['supplier_store'] if s else '', float(s['qty']) if s else 0, float(s['total_cost']) if s else 0,
            (float(r['qty'] - s['qty']) if s else float(r['qty']))
        ])
    # 2026-03-23 固化：订单匹配表中的供应商总成本同样保留原始正负号，负数直接显示。
    write_rows(ws,
        ['平台店铺', '订单号', '平台件数', '订单状态', '匹配结果', '供应商店铺', '供应商件数', '供应商总成本', '件数差'],
        rows, currency_cols={8})

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = True

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    return {
        'output_path': output_path,
        'store_map': store_map,
        'summary_rows': summary_rows,
        'supplier_no_platform_count': len(supplier_no_platform),
        'platform_no_supplier_count': len(platform_no_supplier),
        'refunds_cost_count': len(refunds_cost_rows),
        'validation_rows': validation_rows,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--zip', required=True)
    ap.add_argument('--output', required=True)
    ap.add_argument('--start', default='2026-03-02')
    ap.add_argument('--end', default='2026-03-08')
    ap.add_argument('--rules', default=os.path.join(os.path.dirname(os.path.dirname(__file__)), 'references', 'default_rules.json'))
    ap.add_argument('--no-strict-checks', action='store_true')
    args = ap.parse_args()
    start_date = dt.date.fromisoformat(args.start)
    end_date = dt.date.fromisoformat(args.end)
    references_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'references')
    rules = load_rules(args.rules, start_date=start_date, end_date=end_date, references_dir=references_dir)
    if args.no_strict_checks:
        rules['checks']['strict'] = False
    result = build_report(args.zip, args.output, start_date, end_date, rules)
    print(result)


if __name__ == '__main__':
    main()
